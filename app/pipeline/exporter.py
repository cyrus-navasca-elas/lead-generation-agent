from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.plan import SearchPlan
from app.models.run import RunState
from app.models.score import ScoredCompany

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGH_SCORE_FILL = PatternFill("solid", fgColor="DCFCE7")
MID_SCORE_FILL = PatternFill("solid", fgColor="FEF9C3")


def write_excel(scored: list[ScoredCompany], path: Path) -> None:
    wb = Workbook()

    companies_ws = wb.active
    companies_ws.title = "Companies"
    _write_companies_sheet(companies_ws, scored)

    contacts_ws = wb.create_sheet("Contacts")
    _write_contacts_sheet(contacts_ws, scored)

    enrichment_ws = wb.create_sheet("Enrichment")
    _write_enrichment_sheet(enrichment_ws, scored)

    breakdown_ws = wb.create_sheet("Score Detail")
    _write_breakdown_sheet(breakdown_ws, scored)

    wb.save(path)


def write_json(
    scored: list[ScoredCompany],
    plan: SearchPlan,
    run: RunState,
    path: Path,
) -> None:
    payload: dict[str, Any] = {
        "run": run.model_dump(mode="json"),
        "plan": plan.model_dump(mode="json"),
        "results": [s.model_dump(mode="json") for s in scored],
    }
    path.write_text(json.dumps(payload, indent=2, default=str))


def write_manifest(run: RunState, path: Path) -> None:
    path.write_text(json.dumps(run.model_dump(mode="json"), indent=2, default=str))


def _write_companies_sheet(ws, scored: list[ScoredCompany]) -> None:
    headers = [
        "Company",
        "Source",
        "Industry / Classifications",
        "Employees",
        "Revenue",
        "Location",
        "Website",
        "Technologies",
        "License #",
        "Business Type",
        "Bond Amount",
        "License Status",
        "License Expires",
        "Base Score",
        "Relevance",
        "Total Score",
        "Priority",
        "Confidence",
        "Relevance Reason",
        "Pain Points",
        "Fit Reason",
        "ERP Relevance",
    ]
    _write_headers(ws, headers)

    for row_idx, item in enumerate(scored, start=2):
        s = item.summary
        c = item.company
        rel = item.relevance
        ws.cell(row=row_idx, column=1, value=c.name)
        ws.cell(row=row_idx, column=2, value=c.source)
        ws.cell(row=row_idx, column=3, value=c.industry)
        ws.cell(row=row_idx, column=4, value=c.employee_count)
        ws.cell(row=row_idx, column=5, value=c.revenue)
        ws.cell(row=row_idx, column=6, value=c.location)
        ws.cell(row=row_idx, column=7, value=c.website)
        ws.cell(row=row_idx, column=8, value=", ".join(c.technologies) if c.technologies else None)
        ws.cell(row=row_idx, column=9, value=c.license_number)
        ws.cell(row=row_idx, column=10, value=c.business_type)
        ws.cell(row=row_idx, column=11, value=c.bond_amount)
        ws.cell(row=row_idx, column=12, value=c.license_status)
        ws.cell(
            row=row_idx,
            column=13,
            value=c.license_expiration_date.isoformat() if c.license_expiration_date else None,
        )
        ws.cell(row=row_idx, column=14, value=item.base_score)
        ws.cell(row=row_idx, column=15, value=rel.score if rel else None)
        score_cell = ws.cell(row=row_idx, column=16, value=item.total_score)
        ws.cell(row=row_idx, column=17, value=rel.outreach_priority if rel else "")
        ws.cell(row=row_idx, column=18, value=s.confidence if s else "")
        ws.cell(row=row_idx, column=19, value=rel.reasoning if rel else "")
        ws.cell(row=row_idx, column=20, value="\n".join(s.pain_points) if s else "")
        ws.cell(row=row_idx, column=21, value=s.fit_reason if s else "")
        ws.cell(row=row_idx, column=22, value=s.erp_relevance if s else "")

        if item.total_score >= 70:
            score_cell.fill = HIGH_SCORE_FILL
        elif item.total_score >= 40:
            score_cell.fill = MID_SCORE_FILL

    _autosize(ws, max_width=60)
    _wrap_text_columns(ws, [19, 20, 21, 22])


def _write_contacts_sheet(ws, scored: list[ScoredCompany]) -> None:
    headers = [
        "Company",
        "Name",
        "Title",
        "Department",
        "Email",
        "Phone",
        "Reason to Contact",
    ]
    _write_headers(ws, headers)

    row_idx = 2
    for item in scored:
        reasons = _contact_reason(item)
        for contact in item.contacts:
            ws.cell(row=row_idx, column=1, value=item.company.name)
            ws.cell(row=row_idx, column=2, value=contact.name)
            ws.cell(row=row_idx, column=3, value=contact.title)
            ws.cell(row=row_idx, column=4, value=contact.department)
            ws.cell(row=row_idx, column=5, value=contact.email)
            ws.cell(row=row_idx, column=6, value=contact.phone)
            ws.cell(row=row_idx, column=7, value=reasons)
            row_idx += 1

    _autosize(ws, max_width=50)
    _wrap_text_columns(ws, [7])


def _write_enrichment_sheet(ws, scored: list[ScoredCompany]) -> None:
    headers = [
        "Company",
        "License #",
        "Website",
        "Description",
        "Services",
        "Est. Employees",
        "Years in Business",
        "Web Signals",
        "Contact Emails",
        "Scrape Chars",
        "Confidence",
        "Relevance Score",
        "Outreach Priority",
        "Relevance Reasoning",
    ]
    _write_headers(ws, headers)

    row_idx = 2
    for item in scored:
        p = item.enriched_profile
        r = item.relevance
        if not p and not r:
            continue
        c = item.company
        ws.cell(row=row_idx, column=1, value=c.name)
        ws.cell(row=row_idx, column=2, value=c.license_number)
        ws.cell(row=row_idx, column=3, value=p.website if p else None)
        ws.cell(row=row_idx, column=4, value=p.description if p else None)
        ws.cell(row=row_idx, column=5, value=", ".join(p.services) if p and p.services else None)
        ws.cell(row=row_idx, column=6, value=p.estimated_employees if p else None)
        ws.cell(row=row_idx, column=7, value=p.years_in_business if p else None)
        ws.cell(row=row_idx, column=8, value="\n".join(p.signals) if p and p.signals else None)
        ws.cell(row=row_idx, column=9, value=", ".join(p.contact_emails) if p and p.contact_emails else None)
        ws.cell(row=row_idx, column=10, value=p.scraped_chars if p else None)
        ws.cell(row=row_idx, column=11, value=p.confidence if p else None)
        ws.cell(row=row_idx, column=12, value=r.score if r else None)
        ws.cell(row=row_idx, column=13, value=r.outreach_priority if r else None)
        ws.cell(row=row_idx, column=14, value=r.reasoning if r else None)
        row_idx += 1

    _autosize(ws, max_width=60)
    _wrap_text_columns(ws, [4, 8, 14])


def _write_breakdown_sheet(ws, scored: list[ScoredCompany]) -> None:
    headers = ["Company", "Rule ID", "Rule", "Points", "Total Score"]
    _write_headers(ws, headers)

    row_idx = 2
    for item in scored:
        for b in item.breakdown:
            ws.cell(row=row_idx, column=1, value=item.company.name)
            ws.cell(row=row_idx, column=2, value=b.rule_id)
            ws.cell(row=row_idx, column=3, value=b.label)
            ws.cell(row=row_idx, column=4, value=b.points)
            ws.cell(row=row_idx, column=5, value=item.total_score)
            row_idx += 1

    _autosize(ws, max_width=50)


def _write_headers(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                value = "" if cell.value is None else str(cell.value)
                longest = max(longest, max(len(line) for line in value.splitlines() or [""]))
        ws.column_dimensions[letter].width = min(max_width, max(12, longest + 2))


def _wrap_text_columns(ws, columns: list[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        for col in columns:
            if col <= len(row):
                row[col - 1].alignment = Alignment(wrap_text=True, vertical="top")


def _contact_reason(item: ScoredCompany) -> str:
    if item.summary and item.summary.recommended_contacts:
        return "; ".join(item.summary.recommended_contacts[:3])
    keys = [s.label for s in item.signals[:3]]
    return "; ".join(keys) if keys else "Decision-maker at qualified company."
